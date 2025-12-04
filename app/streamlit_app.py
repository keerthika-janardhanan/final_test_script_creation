# streamlit_app.py
import os
import io
import json
import subprocess
import sys
import signal
import time
import hashlib
import re
import pandas as pd
import streamlit as st
from urllib.parse import urlparse
from hashstore import init_db
from vector_db import VectorDBClient
from test_case_generator import TestCaseGenerator, map_llm_to_template
from ingest_utils import ingest_artifact
from ingest import ingest_jira, ingest_web_site, ingest_ui_crawl, ingest_document
from parse_playwright import parse_playwright_code
from locator_generator import generate_xpath_candidates, to_union_xpath
from recorder_auto_ingest import auto_refine_and_ingest
from recorder_enricher import enrich_recorder_flow, persist_enriched_artifacts
from template_utils import load_excel_template
from trial_spec_adapter import prepare_trial_spec_path, trial_env_overrides
import tempfile
import shutil
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime
# from langchain.chat_models import ChatOpenAI
from codegen_utils import generate_final_script
from executor import run_trial
from agentic_script_agent import (
    AgenticScriptAgent,
    FrameworkProfile,
    initialise_agentic_state,
    interpret_confirmation,
    interpret_feedback,
    interpret_push,
)
from openpyxl import load_workbook

# -------------------------- Constants --------------------------
JSON_FLOW_DIR = os.path.join(os.getcwd(), "app", "saved_flows")
LOCATOR_DIR = os.path.join(os.getcwd(), "app", "./locators")
os.makedirs(JSON_FLOW_DIR, exist_ok=True)
os.makedirs("uploads", exist_ok=True)
FRAMEWORK_CLONE_BASE = Path(os.getcwd()) / "framework_repos"
FRAMEWORK_CLONE_BASE.mkdir(exist_ok=True)

# -------------------------- Initialize DB & Vector --------------------------
init_db()
db = VectorDBClient()
agentic_engine = AgenticScriptAgent()

# -------------------------- Page Config --------------------------
st.set_page_config(page_title="Test Artifact Recorder & Ingest", layout="wide")

# -------------------------- Authentication (demo) --------------------------
if "role" not in st.session_state:
    st.session_state["role"] = "user"

if "agentic_state" not in st.session_state:
    st.session_state.agentic_state = initialise_agentic_state()

st.session_state.setdefault("framework_repo_path", "")
st.session_state.setdefault("framework_branch", "main")
st.session_state.setdefault("framework_commit_message", "Add generated Playwright test")
st.session_state.setdefault("resolved_framework_path", "")

st.sidebar.header("Login (demo)")
st.session_state["role"] = st.sidebar.selectbox(
    "Select role", ["user", "admin"], index=0, key="role_select"
)

st.sidebar.markdown("---")
st.sidebar.subheader("Agentic Script Settings")
st.session_state.framework_repo_path = st.sidebar.text_input(
    "Framework Repo Path",
    value=st.session_state.framework_repo_path,
    key="framework_repo_path_input",
)
st.session_state.framework_branch = st.sidebar.text_input(
    "Git Branch",
    value=st.session_state.framework_branch,
    key="framework_branch_input",
)
st.session_state.framework_commit_message = st.sidebar.text_input(
    "Commit Message",
    value=st.session_state.framework_commit_message,
    key="framework_commit_message_input",
)
st.session_state.setdefault("rec_python_executable", sys.executable)
st.session_state.rec_python_executable = st.sidebar.text_input(
    "Recorder Python Executable",
    value=st.session_state.rec_python_executable,
    help="Optional path to the Python executable used to launch the recorder (defaults to Streamlit's Python).",
    key="rec_python_executable_input",
)

st.title("Test Artifact Recorder & Ingest")

# -------------------------- Helper: Flatten Metadata --------------------------
def flatten_metadata(meta: dict) -> dict:
    """Flatten metadata and remove None values for Chroma compatibility."""
    flat = {}
    for k, v in meta.items():
        if v is None:
            continue
        elif isinstance(v, (dict, list)):
            flat[k] = json.dumps(v, ensure_ascii=False)
        else:
            flat[k] = v
    return flat

def _normalise_keyword(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())

def _extract_locator_info(step: Dict[str, Any]) -> Tuple[str, bool, str]:
    """
    Return (usable_locator, has_locator, raw_locator).
    Treat generic placeholders (body/document/window) as missing.
    """
    raw_locator = ""
    locators = step.get("locators")
    if isinstance(locators, dict):
        for key in ("playwright", "stable", "css", "xpath", "text"):
            candidate = locators.get(key)
            if isinstance(candidate, str) and candidate.strip():
                cleaned = candidate.strip()
                raw_locator = cleaned
                if cleaned.lower() in {"body", "document", "window"}:
                    return "", False, raw_locator
                return cleaned, True, raw_locator
    for key in ("selector", "target", "locator"):
        candidate = step.get(key)
        if isinstance(candidate, str) and candidate.strip():
            cleaned = candidate.strip()
            raw_locator = cleaned
            if cleaned.lower() in {"body", "document", "window"}:
                return "", False, raw_locator
            return cleaned, True, raw_locator
    return "", False, raw_locator

def _step_contains_keyword(step: Dict[str, Any], normalised_keyword: str) -> bool:
    if not normalised_keyword:
        return True
    fields = [
        step.get("action"),
        step.get("navigation"),
        step.get("summary"),
        step.get("name"),
        step.get("label"),
        step.get("expected"),
    ]
    for field in fields:
        if field:
            if normalised_keyword in _normalise_keyword(str(field)):
                return True
    locators = step.get("locators") or {}
    if isinstance(locators, dict):
        for value in locators.values():
            if isinstance(value, str) and normalised_keyword in _normalise_keyword(value):
                return True
    selector = step.get("selector")
    if isinstance(selector, str) and normalised_keyword in _normalise_keyword(selector):
        return True
    return False


def _primary_locator(step: Dict[str, Any]) -> str:
    locators = (step or {}).get("locators") or {}
    for key in ("playwright", "stable", "css", "text", "xpath"):
        candidate = locators.get(key)
        if isinstance(candidate, str) and candidate.strip():
            return candidate.strip()
    navigation = step.get("navigation")
    if isinstance(navigation, str) and navigation.strip():
        return navigation.strip()
    action = step.get("action")
    if isinstance(action, str) and action.strip():
        return action.strip()
    return ""


def format_refined_steps(vector_steps: List[Dict[str, Any]], limit: int = 6) -> str:
    if not vector_steps:
        return ""
    limit = min(limit, len(vector_steps))
    lines = [f"Refined recorder steps (first {limit}):"]
    for step in vector_steps[:limit]:
        summary = step.get("navigation") or step.get("action") or "(step)"
        locator = _primary_locator(step)
        lines.append(f"{step.get('step', '?')}. {summary}")
        if locator:
            lines.append(f"    Locator: {locator}")
        expected = step.get("expected")
        if expected:
            lines.append(f"    Expected: {expected}")
    return "\n".join(lines)


def format_refined_diff(
    vector_steps: List[Dict[str, Any]],
    repo_contents: List[str],
    limit: int = 12,
) -> str:
    if not vector_steps:
        return ""
    repo_blob = "\n".join(repo_contents) if repo_contents else ""
    limit = min(limit, len(vector_steps))
    lines = [f"Repo vs refined recorder steps (first {limit}):"]
    for step in vector_steps[:limit]:
        locator = _primary_locator(step)
        summary = locator or step.get("navigation") or step.get("action") or "(step)"
        change = "REUSED" if locator and locator in repo_blob else "ADDED"
        lines.append(f"{step.get('step', '?')}. {change}: {summary}")
    return "\n".join(lines)

def _filter_steps_with_locators(steps: List[Dict[str, Any]], normalised_keyword: str) -> List[Dict[str, Any]]:
    return _filter_steps_with_locators_ext(steps, normalised_keyword, allow_missing=False)

def _filter_steps_with_locators_ext(
    steps: List[Dict[str, Any]],
    normalised_keyword: str,
    allow_missing: bool = False,
) -> List[Dict[str, Any]]:
    filtered: List[Dict[str, Any]] = []
    for step in steps or []:
        locator, has_locator, raw_locator = _extract_locator_info(step)
        if not has_locator and not allow_missing:
            continue
        if not _step_contains_keyword(step, normalised_keyword):
            continue
        entry = dict(step)
        entry["locator"] = locator or raw_locator or ""
        entry["has_locator"] = has_locator
        filtered.append(entry)
    return filtered

def _format_step_summary(steps: List[Dict[str, Any]], limit: int = 8) -> str:
    lines: List[str] = []
    for idx, step in enumerate(steps[:limit], start=1):
        elements: List[str] = []
        if step.get("action"):
            elements.append(str(step["action"]))
        elif step.get("navigation"):
            elements.append(str(step["navigation"]))
        elif step.get("summary"):
            elements.append(str(step["summary"]))
        label = step.get("label") or step.get("name")
        if label:
            elements.append(str(label))
        locator_value = step.get("locator")
        has_locator = step.get("has_locator", bool(locator_value))
        if locator_value:
            if has_locator:
                elements.append(f"[{locator_value}]")
            else:
                elements.append(f"[locator missing -> {locator_value}]")
        elif not has_locator:
            elements.append("[locator missing]")
        if not elements:
            elements.append("Step details unavailable")
        lines.append(f"{idx}. " + " - ".join(elements))
    if len(steps) > limit:
        lines.append(f"... (+{len(steps) - limit} more steps)")
    return "\n".join(lines)


def _scenario_tokens(text: str) -> List[str]:
    return [token for token in re.findall(r"[a-z0-9]+", _normalise_keyword(text)) if len(token) >= 3]


def find_test_manager_path(framework: FrameworkProfile) -> Optional[Path]:
    direct = framework.root / "testmanager.xlsx"
    if direct.exists():
        return direct
    try:
        candidates = sorted(
            framework.root.glob("**/testmanager.xlsx"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
    except Exception:
        candidates = list(framework.root.glob("**/testmanager.xlsx"))
    return candidates[0] if candidates else None


def persist_uploaded_data_file(framework: FrameworkProfile) -> List[Dict[str, str]]:
    payload = st.session_state.get("pending_data_upload")
    if not payload:
        return []

    messages: List[Dict[str, str]] = []
    try:
        data_dir = framework.additional_dirs.get("data") or (framework.root / "data")
        data_dir.mkdir(parents=True, exist_ok=True)
        safe_name = Path(str(payload.get("name", "data.xlsx"))).name
        data_bytes = payload.get("bytes")
        if isinstance(data_bytes, memoryview):
            data_bytes = data_bytes.tobytes()
        if not isinstance(data_bytes, (bytes, bytearray)):
            raise ValueError("Uploaded data payload missing bytes content.")
        target_path = data_dir / safe_name
        target_path.write_bytes(bytes(data_bytes))
        rel_path = target_path.relative_to(framework.root)
        messages.append(
            {
                "role": "assistant",
                "content": f"Uploaded data file saved to '{rel_path}'.",
                "type": "text",
            }
        )
    except Exception as exc:  # noqa: BLE001
        messages.append(
            {
                "role": "assistant",
                "content": f"Failed to save uploaded data file '{payload.get('name', 'data')}' to data folder: {exc}",
                "type": "text",
            }
        )
    finally:
        st.session_state["pending_data_upload"] = None
    return messages


def derive_default_datasheet_fields(test_case_name: str) -> Dict[str, str]:
    core = re.sub(r"[^a-zA-Z0-9]+", " ", (test_case_name or "").strip()).strip()
    if not core:
        stem = "TestData"
    else:
        parts = [part.capitalize() for part in core.split()]
        stem = "".join(parts) or "TestData"
    return {
        "datasheet": f"{stem}Data.xlsx",
        "reference_id": f"{stem}001",
        "id_name": f"{stem}ID",
    }


def parse_datasheet_message(message: str, defaults: Dict[str, str]) -> Optional[Dict[str, str]]:
    lowered = message.strip().lower()
    if not lowered:
        return None
    if "use default" in lowered or lowered in {"default", "defaults"}:
        return defaults.copy()

    patterns = {
        "datasheet": r"(?:datasheet(?:name)?|sheet)\s*[:=]?\s*([^\s,;]+)",
        "reference_id": r"(?:reference(?:id)?|ref)\s*[:=]?\s*([^\s,;]+)",
        "id_name": r"(?:id\s*name|idname|rowid)\s*[:=]?\s*([^\s,;]+)",
    }
    result = defaults.copy()
    found_any = False
    for key, pattern in patterns.items():
        match = re.search(pattern, message, flags=re.IGNORECASE)
        if match:
            value = match.group(1).strip().strip("'\"")
            if value:
                result[key] = value
                found_any = True
    return result if found_any else None


def update_test_manager_entry(
    framework: FrameworkProfile,
    scenario: str,
    execute_value: str = "Yes",
    create_if_missing: bool = True,
    datasheet: Optional[str] = None,
    reference_id: Optional[str] = None,
    id_name: Optional[str] = None,
) -> Optional[Dict[str, Any]]:
    tm_path = find_test_manager_path(framework)
    if not tm_path:
        return None
    try:
        wb = load_workbook(tm_path)
    except Exception:
        return None

    ws = wb.active
    header_map: Dict[str, int] = {}
    for idx, cell in enumerate(ws[1], start=1):
        value = cell.value
        if value is None:
            continue
        header_map[_normalise_keyword(str(value))] = idx

    def _find_column(*candidates: str) -> Optional[int]:
        for candidate in candidates:
            key = _normalise_keyword(candidate)
            for header_key, col_idx in header_map.items():
                if key == header_key or key in header_key:
                    return col_idx
        return None

    desc_col = _find_column("TestCaseDescription", "Scenario", "Description")
    execute_col = _find_column("Execute", "Run", "Enabled")
    id_col = _find_column("TestCaseID", "ID", "Identifier")
    datasheet_col = _find_column("DatasheetName", "DataSheet", "Data Sheet")
    reference_col = _find_column("ReferenceID", "Reference Id", "Reference")
    idname_col = _find_column("IDName", "IdentifierName", "RowIdentifier")

    if not desc_col or not execute_col:
        return None

    scenario_norm = _normalise_keyword(scenario)
    scenario_tokens = set(_scenario_tokens(scenario))

    matched_row = None
    matched_description = None
    best_row = None
    best_score = 0

    for row_idx in range(2, ws.max_row + 1):
        desc_cell = ws.cell(row=row_idx, column=desc_col)
        description = str(desc_cell.value or "").strip()
        case_id = ""
        if id_col:
            case_id = str(ws.cell(row=row_idx, column=id_col).value or "").strip()
        if not description:
            if not case_id:
                continue
        desc_norm = _normalise_keyword(description)
        desc_tokens = set(_scenario_tokens(description))
        id_norm = _normalise_keyword(case_id)
        id_tokens = set(_scenario_tokens(case_id))

        if scenario_norm and id_norm:
            if scenario_norm in id_norm or id_norm in scenario_norm:
                matched_row = row_idx
                matched_description = case_id or description
                break
        if scenario_norm and desc_norm:
            if scenario_norm in desc_norm or desc_norm in scenario_norm:
                matched_row = row_idx
                matched_description = description or case_id
                break
        combined_tokens = desc_tokens | id_tokens
        if scenario_tokens:
            score = len(scenario_tokens & combined_tokens)
            if score > best_score:
                best_score = score
                best_row = row_idx
                matched_description = description or case_id

    if matched_row is None and best_row is not None and best_score > 0:
        matched_row = best_row

    rel_path = str(tm_path.relative_to(framework.root)).replace("\\", "/")

    if matched_row is not None:
        exec_cell = ws.cell(row=matched_row, column=execute_col)
        previous_value = str(exec_cell.value).strip() if exec_cell.value is not None else ""
        changed = False
        if execute_value is not None and previous_value != execute_value:
            exec_cell.value = execute_value
            changed = True
        if datasheet_col and datasheet is not None:
            cell = ws.cell(row=matched_row, column=datasheet_col)
            if (cell.value or "").strip() != str(datasheet).strip():
                cell.value = datasheet
                changed = True
        if reference_col and reference_id is not None:
            cell = ws.cell(row=matched_row, column=reference_col)
            if (cell.value or "").strip() != str(reference_id).strip():
                cell.value = reference_id
                changed = True
        if idname_col and id_name is not None:
            cell = ws.cell(row=matched_row, column=idname_col)
            if (cell.value or "").strip() != str(id_name).strip():
                cell.value = id_name
                changed = True
        if changed:
            wb.save(tm_path)
            return {
                "path": rel_path,
                "mode": "updated",
                "description": matched_description or scenario,
                "previous": previous_value,
                "execute": execute_value,
            }
        return {
            "path": rel_path,
            "mode": "unchanged",
            "description": matched_description or scenario,
            "previous": previous_value,
            "execute": execute_value,
        }

    if not create_if_missing:
        return None

    new_row = ws.max_row + 1
    scenario_text = scenario.strip()
    ws.cell(row=new_row, column=desc_col).value = scenario_text
    ws.cell(row=new_row, column=execute_col).value = execute_value
    if id_col:
        ws.cell(row=new_row, column=id_col).value = scenario_text
    if datasheet_col and datasheet is not None:
        ws.cell(row=new_row, column=datasheet_col).value = datasheet
    if reference_col and reference_id is not None:
        ws.cell(row=new_row, column=reference_col).value = reference_id
    if idname_col and id_name is not None:
        ws.cell(row=new_row, column=idname_col).value = id_name
    wb.save(tm_path)
    return {
        "path": rel_path,
        "mode": "created",
        "description": scenario_text,
        "previous": "",
        "execute": execute_value,
    }


def register_config_update(
    state: Dict[str, Any],
    responses: List[Dict[str, str]],
    update_info: Optional[Dict[str, Any]],
) -> None:
    if not update_info:
        return
    paths = set(state.get("updated_configs", []))
    paths.add(update_info["path"])
    state["updated_configs"] = list(paths)
    mode = update_info.get("mode")
    description = update_info.get("description") or "scenario"
    execute_value = update_info.get("execute", "Yes")
    if mode == "created":
        responses.append(
            {
                "role": "assistant",
                "content": f"Added '{description}' to testmanager.xlsx and set Execute='{execute_value}'.",
                "type": "text",
            }
        )
    elif mode == "updated":
        prev = update_info.get("previous") or ""
        responses.append(
            {
                "role": "assistant",
                "content": f"Set Execute='{execute_value}' for '{description}' in testmanager.xlsx (was '{prev or 'empty'}').",
                "type": "text",
            }
        )
    elif mode == "unchanged":
        responses.append(
            {
                "role": "assistant",
                "content": f"Test manager already enabled for '{description}'.",
                "type": "text",
            }
        )



def extract_test_ids_from_content(content: str) -> List[str]:
    ids: List[str] = []
    pattern = re.compile(r"\b(?:run|test)\(\s*[\'\"]([^\'\"]+)[\'\"]", re.MULTILINE)
    for match in pattern.finditer(content or ""):
        title = match.group(1).strip()
        if title and title not in ids:
            ids.append(title)
    return ids


def enable_tests_for_ids(framework: FrameworkProfile, test_ids: List[str], state: Dict[str, Any], responses: List[Dict[str, str]]) -> None:
    for test_id in test_ids:
        update_info = update_test_manager_entry(framework, test_id, execute_value="Yes", create_if_missing=True)
        if update_info:
            register_config_update(state, responses, update_info)
        else:
            responses.append({"role": "assistant", "content": f"Could not update testmanager.xlsx for '{test_id}'.", "type": "text"})

def load_refined_flow_steps(keyword: str, allow_missing_locators: bool = False) -> Optional[Dict[str, Any]]:
    flows_dir = Path(os.getcwd()) / "app" / "generated_flows"
    if not flows_dir.exists():
        return None
    normalised_keyword = _normalise_keyword(keyword)
    if not normalised_keyword:
        return None
    candidates = sorted(flows_dir.glob("*.refined.json"), key=lambda p: p.stat().st_mtime, reverse=True)
    for path in candidates:
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        flow_name = str(data.get("flow_name") or path.stem)
        normalised_flow = _normalise_keyword(flow_name)
        normalised_stem = _normalise_keyword(path.stem)
        if normalised_keyword not in normalised_flow and normalised_keyword not in normalised_stem:
            continue
        steps = data.get("steps") or []
        filtered_steps = _filter_steps_with_locators_ext(
            steps,
            normalised_keyword,
            allow_missing=allow_missing_locators,
        )
        if not filtered_steps:
            continue
        return {
            "flow_name": flow_name,
            "path": str(path.relative_to(Path(os.getcwd()))),
            "steps": filtered_steps,
            "updated_at": datetime.fromtimestamp(path.stat().st_mtime),
        }
    return None

def load_latest_refined_flow_details(limit_steps: int = 10) -> Optional[Dict[str, Any]]:
    flows_dir = Path(os.getcwd()) / "app" / "generated_flows"
    if not flows_dir.exists():
        return None
    candidates = sorted(flows_dir.glob("*.refined.json"), key=lambda p: p.stat().st_mtime, reverse=True)
    for path in candidates:
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        steps = data.get("steps") or []
        filtered_steps = _filter_steps_with_locators(steps, _normalise_keyword(data.get("flow_name") or path.stem))
        if not filtered_steps:
            continue
        return {
            "flow_name": data.get("flow_name") or path.stem,
            "path": str(path.relative_to(Path(os.getcwd()))),
            "updated_at": datetime.fromtimestamp(path.stat().st_mtime),
            "steps": filtered_steps[:limit_steps],
            "total_steps": len(filtered_steps),
        }
    return None

def collect_vector_flow_steps(scenario: str) -> List[Dict[str, Any]]:
    try:
        vector_steps = agentic_engine._collect_vector_flow_steps(scenario, top_k=64)  # noqa: SLF001
    except Exception:
        vector_steps = []
    normalised_keyword = _normalise_keyword(scenario)
    filtered: List[Dict[str, Any]] = []
    for step in vector_steps or []:
        summary_text = step.get("summary") or ""
        if normalised_keyword and normalised_keyword not in _normalise_keyword(summary_text):
            continue
        locator = (
            step.get("locator")
            or (step.get("locators") or {}).get("playwright")
            or (step.get("locators") or {}).get("css")
            or (step.get("locators") or {}).get("xpath")
        )
        raw_locator = locator.strip() if isinstance(locator, str) else ""
        has_locator = bool(raw_locator and raw_locator.lower() not in {"body", "document", "window"})
        filtered.append(
            {
                "action": step.get("actionType") or step.get("summary") or "",
                "summary": summary_text,
                "locator": raw_locator,
                "has_locator": has_locator,
            }
        )
    return filtered

def compare_repo_and_refined(scenario: str, framework: FrameworkProfile) -> List[Dict[str, str]]:
    stopword_pattern = re.compile(
        r"\b(compare|difference|diff|between|versus|vs\.?|repo|repository|refined|recorder|flow|steps|with|and|git|show|latest|requested)\b",
        flags=re.IGNORECASE,
    )
    stripped = stopword_pattern.sub(" ", scenario)
    keyword_tokens = [
        token
        for token in re.findall(r"[a-z0-9]+", stripped.lower())
        if len(token) >= 3 and token not in {"with", "and", "git", "show", "latest"}
    ]
    if keyword_tokens:
        scenario = " ".join(keyword_tokens[:6])
    normalised_keyword = _normalise_keyword(scenario)
    repo_assets = agentic_engine.find_existing_framework_assets(scenario, framework)
    matched_repo_assets: List[Dict[str, Any]] = []
    repo_contents: List[Tuple[Path, str]] = []
    if repo_assets:
        for asset in repo_assets:
            path_obj = asset.get("path")
            if not isinstance(path_obj, Path):
                continue
            try:
                content = path_obj.read_text(encoding="utf-8", errors="ignore")
            except Exception:
                continue
            if normalised_keyword:
                normalized_path = _normalise_keyword(str(path_obj))
                normalized_excerpt = _normalise_keyword(content[:20000])
                if (
                    normalised_keyword not in normalized_path
                    and normalised_keyword not in normalized_excerpt
                ):
                    continue
            matched_repo_assets.append(asset)
            repo_contents.append((path_obj, content))

    refined_flow = load_refined_flow_steps(scenario, allow_missing_locators=True)
    refined_steps = refined_flow["steps"] if refined_flow else []
    if not refined_steps:
        refined_steps = collect_vector_flow_steps(scenario)
        if refined_steps:
            refined_flow = {
                "flow_name": scenario,
                "path": "vector_db",
                "updated_at": None,
                "steps": refined_steps,
                "total_steps": len(refined_steps),
            }

    responses: List[Dict[str, str]] = []
    if not matched_repo_assets and not refined_steps:
        responses.append(
            {
                "role": "assistant",
                "content": "No matching repository scripts or refined recorder flows were found for this scenario.",
                "type": "text",
            }
        )
        return responses

    summary_lines: List[str] = []
    if matched_repo_assets:
        summary_lines.append("Repository files considered:")
        for asset in matched_repo_assets:
            path_obj = asset["path"]
            rel = path_obj.relative_to(framework.root)
            summary_lines.append(f"- {rel}")
    else:
        summary_lines.append("Repository scan: no files matched this scenario.")

    if refined_flow:
        flow_header = f"Refined flow source: {refined_flow['flow_name']} ({refined_flow['path']})"
        if refined_flow.get("updated_at"):
            flow_header += f" â€“ updated {refined_flow['updated_at'].strftime('%Y-%m-%d %H:%M:%S')}"
        summary_lines.append("")
        summary_lines.append(flow_header)
        summary_lines.append(_format_step_summary(refined_steps))
    else:
        summary_lines.append("")
        summary_lines.append("No refined recorder flow or vector-ingested steps available for this scenario.")

    missing_locators: List[str] = []
    matched_count = 0
    total_with_locators = 0
    if refined_steps and repo_contents:
        for step in refined_steps:
            locator = step.get("locator") or ""
            has_locator = step.get("has_locator", bool(locator))
            description = step.get("action") or step.get("summary") or step.get("navigation") or locator or "Unnamed step"
            if has_locator and locator:
                total_with_locators += 1
                in_repo = any(locator in content for _, content in repo_contents)
                if in_repo:
                    matched_count += 1
                else:
                    missing_locators.append(f"- {description} [{locator}]")
            else:
                missing_locators.append(f"- {description} [locator missing]")

    summary_lines.append("")
    if refined_steps and repo_contents:
        if total_with_locators:
            summary_lines.append(
                f"Locator coverage: {matched_count}/{total_with_locators} refined steps found in repository files."
            )
        else:
            summary_lines.append("Refined flow contains no actionable locators to compare against the repository.")
        if missing_locators:
            summary_lines.append("Steps needing repository updates:")
            summary_lines.extend(missing_locators[:8])
            if len(missing_locators) > 8:
                summary_lines.append(f"... (+{len(missing_locators) - 8} more)")
    elif refined_steps and not repo_contents:
        summary_lines.append("Refined steps are available, but no repository files matched this scenario.")
    elif repo_contents and not refined_steps:
        summary_lines.append("Repository files exist, but no refined steps are available for comparison.")

    responses.append({"role": "assistant", "content": "\n".join(summary_lines).strip(), "type": "text"})
    return responses

def present_latest_flow_details(limit_steps: int = 10) -> List[Dict[str, str]]:
    latest = load_latest_refined_flow_details(limit_steps=limit_steps)
    if not latest:
        return [
            {
                "role": "assistant",
                "content": "No refined recorder flows have been ingested yet.",
                "type": "text",
            }
        ]
    header = f"Latest refined flow: {latest['flow_name']} ({latest['path']})"
    if latest.get("updated_at"):
        header += f" - updated {latest['updated_at'].strftime('%Y-%m-%d %H:%M:%S')}"
    else:
        header += " - updated time unknown"
    return [
        {"role": "assistant", "content": header, "type": "text"},
        {"role": "assistant", "content": body or "(No grounded steps available)", "type": "text"},
        {"role": "assistant", "content": footer, "type": "text"},
    ]

def execute_push_to_github(state: Dict[str, Any], framework: FrameworkProfile) -> List[Dict[str, str]]:
    responses: List[Dict[str, str]] = []
    payload = state.get("payload")
    if not payload:
        return [{"role": "assistant", "content": "No script payload available to push.", "type": "text"}]

    if state.get("last_trial_status") != "pass":
        return [
            {
                "role": "assistant",
                "content": "Please run the trial and ensure it passes before pushing changes to Git.",
                "type": "text",
            }
        ]

    if not state.get("written_files"):
        try:
            written_paths = agentic_engine.persist_payload(framework, payload)
        except Exception as exc:  # noqa: BLE001
            return [{"role": "assistant", "content": f"Failed to persist files: {exc}", "type": "text"}]
        state["written_files"] = [str(path.relative_to(framework.root)) for path in written_paths]

    config_paths = state.get("updated_configs", [])
    if config_paths:
        state.setdefault("written_files", [])
        for rel in config_paths:
            if rel not in state["written_files"]:
                state["written_files"].append(rel)

    success = agentic_engine.push_changes(
        framework,
        branch=st.session_state.framework_branch,
        commit_msg=st.session_state.framework_commit_message,
    )
    if success:
        state["status"] = "complete"
        state["active"] = False
        files_list = "\n".join(f"- {path}" for path in state.get("written_files", []))
        responses.append(
            {
                "role": "assistant",
                "content": f"Changes pushed successfully. Files:\n{files_list}",
                "type": "text",
            }
        )
    else:
        responses.append(
            {
                "role": "assistant",
                    "content": "Git push failed. Please check repository permissions and try again.",
                    "type": "text",
                }
            )
    return responses


def _build_playwright_command(spec_path: Path | str, headed: bool = False) -> List[str]:
    if isinstance(spec_path, Path):
        spec_arg = spec_path.as_posix()
    else:
        spec_arg = str(spec_path).replace("\\", "/")
    base_cmd = ["npx", "playwright", "test", spec_arg]
    if os.name == "nt":
        base_cmd[0] = "npx.cmd"
    # Trial runs default to a single worker to avoid multiple browsers competing for shared credentials/session.
    base_cmd.append("--workers=1")
    if headed:
        base_cmd.append("--headed")
    return base_cmd


def run_spec_file(spec_path: Path, repo_root: Path, headed: bool = False) -> Tuple[bool, str]:
    absolute_spec = spec_path if spec_path.is_absolute() else (repo_root / spec_path).resolve()
    prepared_path, cleanup = prepare_trial_spec_path(absolute_spec, repo_root)
    try:
        # Resolve path for CLI argument
        try:
            relative_prepared = prepared_path.relative_to(repo_root)
            spec_arg: Path | str = relative_prepared
        except ValueError:
            spec_arg = prepared_path

        cmd = _build_playwright_command(spec_arg, headed=headed)
        env_vars = os.environ.copy()
        try_env = trial_env_overrides(repo_root, spec_path=prepared_path)
        if try_env:
            env_vars.update(try_env)

        try:
            result = subprocess.run(
                cmd,
                cwd=str(repo_root),
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                env=env_vars,
            )
            success = result.returncode == 0
            stdout = (result.stdout or "").strip()
            stderr = (result.stderr or "").strip()
            logs = "\n".join(part for part in [stdout, stderr] if part).strip()
            return success, logs or "(no output)"
        except FileNotFoundError:
            fallback_cmd = (
                repo_root / "node_modules" / ".bin" / ("playwright.cmd" if os.name == "nt" else "playwright")
            )
            if fallback_cmd.exists():
                spec_cli_arg = spec_arg.as_posix() if isinstance(spec_arg, Path) else spec_arg
                try:
                    result = subprocess.run(
                        [str(fallback_cmd), "test", spec_cli_arg] + (["--headed"] if headed else []),
                        cwd=str(repo_root),
                        capture_output=True,
                        text=True,
                        encoding="utf-8",
                        errors="replace",
                        env=env_vars,
                    )
                    success = result.returncode == 0
                    stdout = (result.stdout or "").strip()
                    stderr = (result.stderr or "").strip()
                    logs = "\n".join(part for part in [stdout, stderr] if part).strip()
                    return success, logs or "(no output)"
                except Exception as exc:  # noqa: BLE001
                    return False, str(exc)
            return False, (
                "Playwright CLI not found. Ensure Node.js is in PATH and run "
                "`npm install` followed by `npx playwright install` in the repository root."
            )
    except Exception as exc:  # noqa: BLE001
        return False, str(exc)
    finally:
        if cleanup:
            cleanup()


def execute_trial_run(
    state: Dict[str, Any],
    framework: FrameworkProfile,
    scenario_hint: str = "",
    headed: bool = False,
) -> List[Dict[str, str]]:
    # Deprecated: Streamlit trial path is disabled. Use FastAPI route instead.
    state["last_trial_status"] = "skipped"
    state["last_trial_mode"] = "headless"
    return [
        {
            "role": "assistant",
            "content": (
                "Trial run via Streamlit is disabled. "
                "Please call the FastAPI endpoint POST /agentic/trial-run-existing "
                "to execute trials (supports unskip and parallel ReferenceIDs)."
            ),
            "type": "text",
        }
    ]


# -------------------------- Admin Panel --------------------------
if st.session_state["role"] == "admin":
    st.header("Admin: Ingest & Manage")

    # ---------------- Jira ----------------
    st.subheader("Jira Ingestion")
    jql_input = st.text_input(
        "Jira JQL",
        value="project=GEN_AI_PROJECT ORDER BY created DESC",
        key="jira_jql_input"
    )
    if st.button("Fetch & Ingest Jira", key="btn_ingest_jira"):
        try:
            results = ingest_jira(jql_input)
            st.success(f"Jira stories ingested: {len(results)} issues âœ…")
        except Exception as e:
            st.error(f"Jira ingestion failed: {e}")

    # ---------------- Website ----------------
    st.subheader("Website Ingestion")
    url = st.text_input(
        "Website URL",
        value="https://docs.oracle.com/en/cloud/saas/index.html",
        key="website_url_input"
    )
    max_depth = st.number_input(
        "Max Depth", min_value=1, max_value=5, value=2, key="website_max_depth"
    )
    if st.button("Fetch & Ingest Website", key="btn_ingest_website"):
        if url.strip():
            try:
                with st.spinner(f"Crawling {url} up to depth {max_depth}..."):
                    results = ingest_web_site(url, max_depth)
                st.success(f"Website ingestion finished: {len(results)} docs added âœ…")
            except Exception as e:
                st.error(f"Website ingestion failed: {e}")
        else:
            st.warning("Please enter a valid URL")

    # ---------------- Document Ingestion ----------------
    st.subheader("Document Ingestion")
    uploaded_files = st.file_uploader(
        "Upload documents (PDF, DOCX, TXT)",
        type=["pdf", "docx", "doc", "txt"],
        accept_multiple_files=True,
        key="doc_uploader"
    )

    if uploaded_files:
        if st.button("Ingest Uploaded Documents", key="btn_ingest_docs"):
            all_results = []
            try:
                for uploaded_file in uploaded_files:
                    temp_path = os.path.join("uploads", uploaded_file.name)
                    with open(temp_path, "wb") as f:
                        f.write(uploaded_file.getbuffer())
                    with st.spinner(f"Ingesting {uploaded_file.name}..."):
                        results = ingest_document(temp_path)
                        all_results.extend(results)
                st.success(f"Document ingestion finished: {len(all_results)} docs added âœ…")
            except Exception as e:
                st.error(f"Document ingestion failed: {e}")

    # ---------------- UI Crawl ----------------
    st.subheader("UI Crawl Ingestion")
    crawl_file = st.file_uploader(
        "Upload crawl JSON", type=["json"], key="crawl_file_uploader"
    )
    if st.button("Ingest UI Crawl", key="btn_ingest_ui_crawl"):
        if crawl_file:
            path = f"./uploads/{crawl_file.name}"
            with open(path, "wb") as f:
                f.write(crawl_file.getbuffer())
            try:
                results = ingest_ui_crawl(path)
                st.success(f"UI Crawl ingested: {len(results)} flows âœ…")
            except Exception as e:
                st.error(f"UI Crawl ingestion failed: {e}")
        else:
            st.warning("Please upload a crawl JSON file")

    # ---------------- Delete Management ----------------
    st.subheader("Manage Vector DB Documents")
    delete_mode = st.radio(
        "Choose delete mode", ["By ID", "By Source"], key="delete_mode_radio"
    )
    
    if delete_mode == "By ID":
        doc_id_input = st.text_input("Enter Document ID to delete", key="delete_doc_id")
        if st.button("ðŸ—‘ï¸ Delete Document by ID", key="btn_delete_by_id"):
            if doc_id_input.strip():
                try:
                    db.delete_document(doc_id_input.strip())
                    st.success(f"Document '{doc_id_input}' deleted successfully âœ…")
                except Exception as e:
                    st.error(f"Failed to delete document: {e}")
            else:
                st.warning("Please enter a valid Document ID.")
    
    elif delete_mode == "By Source":
        source_input = st.text_input("Enter Source (e.g. 'jira', 'ui_flow')", key="delete_source_input")
        if st.button("ðŸ—‘ï¸ Delete All Documents by Source", key="btn_delete_by_source"):
            if source_input.strip():
                try:
                    db.delete_by_source(source_input.strip())
                    st.success(f"All documents from source '{source_input}' deleted âœ…")
                except Exception as e:
                    st.error(f"Failed to delete by source: {e}")
            else:
                st.warning("Please enter a valid source name.")

    # ---------------- Show Existing Docs ----------------
    if st.checkbox("ðŸ“‹ Show Existing Docs with Pagination", key="show_docs_checkbox"):
        try:
            all_docs = db.list_all(limit=1000)
            if all_docs:
                page_size = st.number_input("Docs per page", min_value=5, max_value=100, value=20, key="docs_page_size")
                total_pages = (len(all_docs) + page_size - 1) // page_size
                current_page = st.number_input("Page", min_value=1, max_value=total_pages, value=1, key="docs_page_num")

                start_idx = (current_page - 1) * page_size
                end_idx = start_idx + page_size
                page_docs = all_docs[start_idx:end_idx]

                df = pd.DataFrame(page_docs)
                st.dataframe(df)

                st.write(f"Showing page {current_page} of {total_pages}")
            else:
                st.info("No documents found in Vector DB.")
        except Exception as e:
            st.error(f"Failed to fetch documents: {e}")

# -------------------------- Playwright Recorder Panel --------------------------
st.header("Playwright Recorder â†’ Vector DB Ingestion")
flow_name = st.text_input("Flow Name", "playwright-recorded-flow")
record_url = st.text_input("URL to Record", "https://example.com")

if "record_proc" not in st.session_state:
    st.session_state["record_proc"] = None
if "record_session_dir" not in st.session_state:
    st.session_state["record_session_dir"] = None
if "record_metadata" not in st.session_state:
    st.session_state["record_metadata"] = None
if "record_manual_out_path" not in st.session_state:
    st.session_state["record_manual_out_path"] = None
if "record_manual_log" not in st.session_state:
    st.session_state["record_manual_log"] = ""
if "record_session_listing" not in st.session_state:
    st.session_state["record_session_listing"] = None


def _load_recorder_metadata(session_dir: str, attempts: int = 15, delay: float = 0.5) -> Optional[dict]:
    session_path = Path(session_dir)
    metadata_path = session_path / "metadata.json"
    for _ in range(attempts):
        if metadata_path.exists():
            try:
                return json.loads(metadata_path.read_text(encoding="utf-8"))
            except json.JSONDecodeError:
                time.sleep(delay)
                continue
        time.sleep(delay)
    return None


def _scan_session_directory(session_dir: str) -> Dict[str, Any]:
    session_path = Path(session_dir)
    summary: Dict[str, Any] = {
        "exists": session_path.exists(),
        "top_level": [],
        "dom_files": 0,
        "screenshot_files": 0,
    }
    if not session_path.exists():
        return summary
    try:
        summary["top_level"] = sorted(p.name for p in session_path.iterdir())
    except Exception:
        summary["top_level"] = []
    dom_dir = session_path / "dom"
    if dom_dir.exists():
        summary["dom_files"] = len(list(dom_dir.glob("*.html")))
    shots_dir = session_path / "screenshots"
    if shots_dir.exists():
        summary["screenshot_files"] = len(list(shots_dir.glob("*.png")))
    return summary


def _normalize_record_url(raw_url: str) -> Tuple[Optional[str], Optional[str]]:
    """Return a sanitized http(s) URL suitable for the recorder command."""

    url = (raw_url or "").strip()
    if not url:
        return None, "Please enter a URL to record."

    if "://" not in url:
        url = f"https://{url}"

    parsed = urlparse(url)
    if not parsed.netloc:
        return None, "The recording URL must include a valid host name."

    if parsed.scheme.lower() not in {"http", "https"}:
        return None, "Only http and https URLs are supported for recording."

    normalized = parsed.geturl()
    return normalized, None


def _validate_recorder_runtime(python_exec: str) -> Optional[str]:
    """Ensure the selected interpreter can launch the recorder."""

    try:
        result = subprocess.run(
            [python_exec, "-c", "import playwright"],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            check=False,
        )
    except FileNotFoundError as exc:
        return f"Recorder Python executable not found: {exc}"
    except Exception as exc:  # noqa: BLE001
        return f"Failed to validate recorder runtime ({python_exec}): {exc}"

    if result.returncode != 0:
        stderr = (result.stderr or "").strip()
        stdout = (result.stdout or "").strip()
        details = stderr or stdout or "Playwright import failed for an unknown reason."
        instructions = (
            "Ensure Playwright is installed for this interpreter by running:\n"
            f"{python_exec} -m pip install playwright\n"
            f"{python_exec} -m playwright install chromium"
        )
        return f"Recorder dependencies are missing: {details}\n\n{instructions}"

    return None


def _finalize_recorder_session() -> None:
    session_dir = st.session_state.get("record_session_dir")
    if not session_dir:
        return
    listing = _scan_session_directory(session_dir)
    st.session_state["record_session_listing"] = listing
    metadata = _load_recorder_metadata(session_dir)
    if metadata:
        st.session_state["record_metadata"] = metadata
        missing_parts = []
        options = metadata.get("options", {})
        artifacts = metadata.get("artifacts", {})
        if options.get("captureDom") and not listing.get("dom_files"):
            missing_parts.append("DOM snapshots")
        if options.get("captureScreenshots") and not listing.get("screenshot_files"):
            missing_parts.append("screenshots")
        if options.get("recordTrace") and not artifacts.get("trace"):
            missing_parts.append("trace.zip")
        if options.get("recordHar") and not artifacts.get("har"):
            missing_parts.append("network.har")
        if missing_parts:
            st.warning(
                "Recorder metadata loaded but some expected artefacts appear to be missing: "
                + ", ".join(missing_parts)
            )

        auto_state_key = f"auto_ingest::{session_dir}"
        if auto_state_key not in st.session_state:
            try:
                auto_result = auto_refine_and_ingest(session_dir, metadata)
                st.session_state[auto_state_key] = {
                    "status": "success",
                    "result": auto_result,
                }
            except Exception as exc:  # noqa: BLE001
                st.session_state[auto_state_key] = {
                    "status": "error",
                    "error": str(exc),
                }

        auto_state = st.session_state.get(auto_state_key)
        if auto_state:
            if auto_state["status"] == "success":
                result = auto_state["result"]
                stats = result.get("ingest_stats") or {}
                added = stats.get("added")
                if added is not None:
                    message = f"Refined flow ingested into vector DB ({added} steps)."
                else:
                    message = "Refined flow saved locally."
                st.success(f"{message} Saved to `{result['refined_path']}`.")
            else:
                st.warning(f"Automatic refinement failed: {auto_state['error']}")
    else:
        existing = listing.get("top_level", []) if listing else []
        st.warning(
            "Recorder stopped but metadata.json is not available. "
            "Observed session directory contents: "
            + (", ".join(existing) if existing else "<empty>")
        )

proc = st.session_state.get("record_proc")
if proc and proc.poll() is not None:
    st.session_state["record_proc"] = None
    _finalize_recorder_session()

if "rec_output_dir" not in st.session_state:
    st.session_state["rec_output_dir"] = "recordings"
if "rec_capture_dom" not in st.session_state:
    st.session_state["rec_capture_dom"] = False
if "rec_capture_screens" not in st.session_state:
    st.session_state["rec_capture_screens"] = False
if "rec_capture_trace" not in st.session_state:
    st.session_state["rec_capture_trace"] = True
if "rec_capture_har" not in st.session_state:
    st.session_state["rec_capture_har"] = True
if "rec_ignore_https" not in st.session_state:
    st.session_state["rec_ignore_https"] = False
if "rec_disable_gpu" not in st.session_state:
    st.session_state["rec_disable_gpu"] = False
if "rec_proxy" not in st.session_state:
    st.session_state["rec_proxy"] = ""
if "rec_timeout" not in st.session_state:
    st.session_state["rec_timeout"] = 0

st.text_input("Recording Output Directory", key="rec_output_dir")
opt_cols = st.columns(4)
with opt_cols[0]:
    st.checkbox("Capture DOM Snapshots", key="rec_capture_dom")
with opt_cols[1]:
    st.checkbox("Capture Screenshots", key="rec_capture_screens")
with opt_cols[2]:
    st.checkbox("Capture Playwright Trace", key="rec_capture_trace")
with opt_cols[3]:
    st.checkbox("Capture HAR", key="rec_capture_har")

st.checkbox(
    "Ignore HTTPS Errors",
    key="rec_ignore_https",
    help="Bypass TLS certificate validation (needed for some internal test environments).",
)

gpu_col, proxy_col = st.columns([1, 3])
with gpu_col:
    st.checkbox("Disable GPU", key="rec_disable_gpu", help="Pass GPU-disabling flags to Chromium to avoid blank rendering on some drivers.")
with proxy_col:
    st.text_input("Proxy server (optional)", key="rec_proxy", help="e.g. http://proxy.mycorp:3128")

timeout_col, _ = st.columns([1, 3])
with timeout_col:
    st.number_input("Auto-stop after (seconds)", min_value=0, max_value=3600, step=60, key="rec_timeout")

status_placeholder = st.empty()
if st.session_state.get("record_proc"):
    status_placeholder.info("Recorder is running. Complete your browser actions and stop when finished.")

col1, col2 = st.columns(2)
with col1:
    if st.button("Start Recording") and not st.session_state["record_proc"]:
        normalized_url, url_error = _normalize_record_url(record_url)
        if url_error:
            st.error(url_error)
        else:
            session_name = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            output_root = Path(st.session_state["rec_output_dir"]).expanduser().resolve()
            output_root.mkdir(parents=True, exist_ok=True)
            session_dir = output_root / session_name
            # Create session folder early and persist a human-friendly flow label for better matching later
            try:
                session_dir.mkdir(parents=True, exist_ok=True)
                (session_dir / "flow_name.txt").write_text(flow_name or "", encoding="utf-8")
            except Exception:
                pass
            python_exec = st.session_state.get("rec_python_executable") or sys.executable

            runtime_error = _validate_recorder_runtime(python_exec)
            if runtime_error:
                st.error(runtime_error)
                st.session_state["record_manual_log"] = runtime_error
            else:
                cmd: List[str] = [
                    python_exec,
                    "-m",
                    "app.run_playwright_recorder_v2",
                    "--url",
                    normalized_url,
                    "--output-dir",
                    str(output_root),
                    "--session-name",
                    session_name,
                ]
                if not st.session_state["rec_capture_trace"]:
                    cmd.append("--no-trace")
                if not st.session_state["rec_capture_har"]:
                    cmd.append("--no-har")
                if st.session_state["rec_capture_dom"]:
                    cmd.append("--capture-dom")
                if st.session_state["rec_capture_screens"]:
                    cmd.append("--capture-screenshots")
                if st.session_state["rec_ignore_https"]:
                    cmd.append("--ignore-https-errors")
                if st.session_state.get("rec_disable_gpu"):
                    cmd.append("--disable-gpu")
                if st.session_state.get("rec_proxy"):
                    cmd.extend(["--proxy", st.session_state["rec_proxy"]])
                if st.session_state["rec_timeout"]:
                    cmd.extend(["--timeout", str(int(st.session_state["rec_timeout"]))])

                creationflags = 0
                if os.name == "nt" and hasattr(subprocess, "CREATE_NEW_PROCESS_GROUP"):
                    creationflags = subprocess.CREATE_NEW_PROCESS_GROUP
                try:
                    proc = subprocess.Popen(cmd, creationflags=creationflags)
                    st.session_state["record_proc"] = proc
                    st.session_state["record_session_dir"] = str(session_dir)
                    st.session_state["record_metadata"] = None
                    st.session_state["record_manual_out_path"] = None
                    st.session_state["record_manual_log"] = ""
                    st.success(
                        f"Recorder started. A browser window should open. Session artefacts will appear in `{session_dir}`."
                    )
                except FileNotFoundError as exc:
                    st.error(f"Failed to launch recorder: {exc}")
                except Exception as exc:  # noqa: BLE001
                    st.error(f"Unexpected error launching recorder: {exc}")

with col2:
    if st.button("Stop Recording") and st.session_state["record_proc"]:
        proc = st.session_state["record_proc"]
        try:
            if os.name == "nt" and hasattr(signal, "CTRL_BREAK_EVENT"):
                proc.send_signal(signal.CTRL_BREAK_EVENT)
            else:
                proc.send_signal(signal.SIGINT)
            proc.wait(timeout=15)
        except subprocess.TimeoutExpired:
            proc.terminate()
            try:
                proc.wait(timeout=5)
            except subprocess.TimeoutExpired:
                proc.kill()
        finally:
            st.session_state["record_proc"] = None
            time.sleep(2)
            _finalize_recorder_session()
        st.info("Recorder stopped. Review captured metadata below.")

session_dir = st.session_state.get("record_session_dir")
session_listing = st.session_state.get("record_session_listing")
metadata = st.session_state.get("record_metadata")
if session_dir and session_listing:
    listing_lines = [
        f"metadata.json: {'present' if 'metadata.json' in session_listing.get('top_level', []) else 'missing'}",
        f"DOM snapshots: {session_listing.get('dom_files', 0)} file(s)",
        f"Screenshots: {session_listing.get('screenshot_files', 0)} file(s)",
    ]
    st.markdown("###### Session Directory Snapshot")
    st.code("\n".join(listing_lines))

if session_dir and metadata:
    session_path = Path(session_dir)
    actions = metadata.get("actions", [])
    st.success(
        f"Session `{session_path.name}` captured {len(actions)} actions "
        f"(HAR={'yes' if metadata['options'].get('recordHar') else 'no'}, "
        f"Trace={'yes' if metadata['options'].get('recordTrace') else 'no'})."
    )

    preview_rows = []
    for action in actions:
        element = action.get("element") or {}
        preview_rows.append(
            {
                "Action ID": action.get("actionId"),
                "Action": action.get("action"),
                "Element": element.get("tagName"),
                "Role": element.get("role"),
                "Name / Label": element.get("ariaLabel") or element.get("name") or element.get("text"),
                "Stable Selector": element.get("stableSelector"),
                "Quadrant": (action.get("boundingBox") or {}).get("quadrant", ""),
            }
        )
    if preview_rows:
        st.dataframe(pd.DataFrame(preview_rows), width="stretch", hide_index=True)

    st.markdown("##### Recorder Artefacts")
    metadata_path = session_path / "metadata.json"
    if metadata_path.exists():
        with open(metadata_path, "rb") as fh:
            st.download_button(
                "Download metadata.json",
                data=fh.read(),
                file_name=f"{session_path.name}_metadata.json",
                mime="application/json",
                key="download_metadata_json",
            )
    artifacts = metadata.get("artifacts", {})
    for label, rel_path in artifacts.items():
        if not rel_path:
            continue
        file_path = session_path / rel_path
        if file_path.exists():
            with open(file_path, "rb") as fh:
                st.download_button(
                    f"Download {label}",
                    data=fh.read(),
                    file_name=f"{session_path.name}_{Path(rel_path).name}",
                    key=f"download_{label}",
                )

    st.markdown("##### Generate Manual Test Cases from Recording")
    trace_rel = artifacts.get("trace")
    recording_source = session_path / trace_rel if trace_rel else None
    if recording_source and recording_source.exists():
        template_upload = st.file_uploader(
            "Upload Excel template", type=["xlsx"], key="rec_manual_template_uploader"
        )
        if st.button("Generate manual_from_recording.xlsx", key="btn_manual_from_recording"):
            if not template_upload:
                st.warning("Please upload an Excel template before generating manual test cases.")
            else:
                temp_dir = Path(tempfile.mkdtemp(prefix="manual_from_recording_"))
                template_path = temp_dir / template_upload.name
                template_path.write_bytes(template_upload.getvalue())
                out_path = temp_dir / f"manual_from_{session_path.name}.xlsx"
                cmd = [
                    sys.executable,
                    "manual_from_recording.py",
                    "--template",
                    str(template_path),
                    "--recording",
                    str(recording_source),
                    "--out",
                    str(out_path),
                ]
                dom_dir = session_path / "dom"
                if metadata["options"].get("captureDom") and dom_dir.exists():
                    cmd.extend(["--dom", str(dom_dir)])
                try:
                    result = subprocess.run(
                        cmd,
                        capture_output=True,
                        text=True,
                        encoding="utf-8",
                        errors="replace",
                        check=True,
                    )
                    st.session_state["record_manual_out_path"] = str(out_path)
                    st.session_state["record_manual_log"] = (result.stdout or "") + (result.stderr or "")
                    st.success("Manual workbook generated successfully.")
                except subprocess.CalledProcessError as exc:
                    st.session_state["record_manual_log"] = (exc.stdout or "") + (exc.stderr or "")
                    st.error(f"manual_from_recording.py failed (exit {exc.returncode}). See logs below.")
        if st.session_state.get("record_manual_log"):
            st.code(st.session_state["record_manual_log"], language="bash")
        manual_out_path = st.session_state.get("record_manual_out_path")
        if manual_out_path and Path(manual_out_path).exists():
            with open(manual_out_path, "rb") as fh:
                st.download_button(
                    "Download manual test cases workbook",
                    data=fh.read(),
                    file_name=Path(manual_out_path).name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_manual_workbook",
                )
    else:
        st.info(
            "Trace artefact not found. Ensure trace capture is enabled or rerun the recorder with the default settings."
        )

# Paste TS code
st.markdown("### Paste Playwright TS Codegen Output")
ts_code = st.text_area("Paste code here...", height=300, key="ts_code_input")

if st.button("ðŸ“¥ Convert, Ingest & Generate Locators", key="btn_convert_ingest") and ts_code.strip():
    try:
        # 1ï¸âƒ£ Parse TS Code â†’ Steps
        steps = parse_playwright_code(ts_code)

        # 2ï¸âƒ£ Save recorder flow JSON locally
        artifact = {"flow_name": flow_name, "source": "playwright", "steps": steps}
        json_path = os.path.join(JSON_FLOW_DIR, f"{flow_name}.json")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(artifact, f, indent=4)

        # 3ï¸âƒ£ Generate TypeScript locator file
        locator_file = os.path.join(LOCATOR_DIR, f"{flow_name}.ts")
        with open(locator_file, "w", encoding="utf-8") as f:
            f.write("export const Locators = {\n")
            for i, step in enumerate(steps):
                if "selector" in step:
                    locator_name = f"step{i+1}_{step['action']}".replace("-", "_")
                    cands = generate_xpath_candidates(step["selector"])  # list[str]
                    union_xpath = to_union_xpath(cands)
                    xpath_value = ("xpath=(" + union_xpath + ")").replace('"', '\\"')
                    f.write(f'  {locator_name}: "{xpath_value}",\n')
            f.write("};\n")

        table_rows, sidecar = enrich_recorder_flow(flow_name, steps)
        enriched_paths = persist_enriched_artifacts(flow_name, table_rows, sidecar)

        st.success(
            f"âœ… Flow '{flow_name}' stored locally. Generated locators and enriched scenario artifacts (cache key: {enriched_paths['cache_key']})."
        )
        st.code(open(locator_file).read(), language="typescript")
        st.json(artifact)

        scenario_df = pd.DataFrame(
            table_rows,
            columns=["sl", "Action", "Navigation Steps", "Key Data Element Examples", "Expected Results"],
        )
        st.dataframe(scenario_df, hide_index=True)

        with open(enriched_paths["csv_path"], "rb") as f:
            st.download_button(
                label="ðŸ“¥ Download Scenario CSV",
                data=f.read(),
                file_name=f"{enriched_paths['cache_key']}.csv",
                mime="text/csv",
                key="download_scenario_csv",
            )

        if enriched_paths.get("xlsx_path"):
            with open(enriched_paths["xlsx_path"], "rb") as f:
                st.download_button(
                    label="ðŸ“¥ Download Scenario XLSX",
                    data=f.read(),
                    file_name=f"{enriched_paths['cache_key']}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_scenario_xlsx",
                )

        with open(enriched_paths["json_path"], "rb") as f:
            st.download_button(
                label="ðŸ“¥ Download Sidecar JSON",
                data=f.read(),
                file_name=f"{enriched_paths['cache_key']}.json",
                mime="application/json",
                key="download_scenario_json",
            )

        low_stability_targets = [
            idx + 1
            for idx, step_payload in enumerate(sidecar)
            if any(target.get("stability_score", 0) < 0.6 for target in step_payload.get("targets", []))
        ]
        summary_message = f"Scenario contains {enriched_paths['step_count']} rows."
        if low_stability_targets:
            summary_message += f" Low-stability selectors flagged at steps: {', '.join(map(str, low_stability_targets))}."
        st.info(summary_message)

    except Exception as e:
        st.error(f"âŒ Failed to process recording: {e}")

# -------------------------- Test Case Generator Panel --------------------------
st.markdown("---")
st.subheader("Generate Test Cases from Jira / Keywords / Stories")
jira_llm_only = st.checkbox("LLM-only (skip deterministic injection)", value=False, help="When enabled, the generator will rely solely on the LLM output without injecting deterministic steps from recorder/refined flows.")
jira_input = st.text_area("Paste Jira story, description, or keywords", key="jira_input_area")
template_file = st.file_uploader(
    "Upload Template File (JSON / Excel / Text / Doc)",
    type=["json","xlsx","xls","txt","doc","docx"],
    key="template_file_uploader"
)

if st.button("Generate & Download Test Cases", key="btn_generate_tc") and jira_input.strip():
    try:
        tcg = TestCaseGenerator(db)
        results = tcg.generate_test_cases(jira_input.strip(), llm_only=jira_llm_only)
        if template_file:
            ext = os.path.splitext(template_file.name)[1].lower()
            if ext in [".xlsx", ".xls"]:
                template_df = load_excel_template(template_file)
                df = map_llm_to_template(results, template_df)
            else:
                df = pd.DataFrame(results)
        else:
            df = pd.DataFrame(results)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            df.to_excel(writer, index=False, sheet_name="TestCases")
        output.seek(0)
        st.download_button(
            label="ðŸ“¥ Download Test Cases as Excel",
            data=output,
            file_name="test_cases.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="btn_download_tc"
        )
    except Exception as e:
        st.error(f"Failed to generate test cases: {e}")

# -------------------------- Manual Table (Markdown) Generator --------------------------
st.markdown("---")
st.subheader("Generate Manual Table (Markdown)")
mt_story = st.text_area(
    "Scenario / Story for manual table",
    key="mt_story_input",
    help="Provide a brief scenario title or paste a story. We'll use refined recorder steps and vector context if available.",
)
mt_query_col, mt_scope_col = st.columns(2)
with mt_query_col:
    mt_query = st.text_input(
        "Retrieval hint (optional)",
        value="",
        key="mt_query_input",
        help="Keywords to help retrieve related flows from the vector DB.",
    )
with mt_scope_col:
    mt_scope = st.text_input(
        "Scope filter (optional)",
        value="",
        key="mt_scope_input",
        help="Optional narrowing hint like 'supplier creation only'.",
    )

mt_cols = st.columns([1, 1, 2])
with mt_cols[0]:
    gen_mt = st.button("Generate Manual Table", key="btn_generate_manual_table")
with mt_cols[1]:
    dl_mt = st.button("Download as .md", key="btn_download_manual_table")

if gen_mt and mt_story.strip():
    try:
        tcg = TestCaseGenerator(db)
        md = tcg.generate_manual_table(mt_story.strip(), db_query=mt_query.strip() or None, scope=mt_scope.strip() or None)
        st.session_state["_last_manual_table_md"] = md
        st.markdown(md)
    except Exception as e:
        st.error(f"Failed to generate manual table: {e}")

if dl_mt:
    md_text = st.session_state.get("_last_manual_table_md", "")
    if not md_text:
        st.warning("No manual table generated yet. Click 'Generate Manual Table' first.")
    else:
        st.download_button(
            label="ðŸ“¥ Download Manual Table",
            data=md_text.encode("utf-8"),
            file_name="manual_test_table.md",
            mime="text/markdown",
            key="download_manual_table_btn",
        )

# -------------------------- Test Script Generator Panel --------------------------
# st.title("AI-Powered Test Script Generator")
# test_case_id = st.text_input("Enter Test Case ID", key="test_case_id_input")

# if st.button("Generate & Run", key="btn_generate_run"):
#     orch = TestScriptOrchestrator()
#     script, success, logs = orch.generate_and_run(test_case_id)

#     st.subheader("Generated Test Script")
#     st.code(script, language="typescript")

#     st.subheader("Execution Result")
#     if success:
#         st.success("âœ… Passed â€“ Script ingested into Vector DB")
#     else:
#         st.error("âŒ Failed â€“ Script NOT ingested")
#     st.text(logs)

# -------------------------- Repo Scaffold Ingestion --------------------------
def ingest_parsed_scaffold(parsed_json):
    """Ingest TS repo scaffold into Vector DB."""
    for module in parsed_json.get("modules", []):
        doc_id = module.get("id") or module.get("name")
        content = json.dumps(module, indent=2)
        metadata = {
            "type": "repo_scaffold",
            "module": module.get("name"),
        }
        db.add_document(source="repo_scaffold", doc_id=doc_id, content=content, metadata=metadata)

st.subheader("Pull Git Repo & Ingest Scaffold")

repo_url = st.text_input(
    "Git Repo URL",
    "https://github.com/keerthika-janardhanan/oracle_erp.git",
    key="repo_url_input2"
)
branch = st.text_input("Branch", "main", key="branch_input2")

def pull_and_ingest_repo(repo_url, branch):
    tmp_dir = tempfile.mkdtemp(prefix="repo_clone_")
    try:
        subprocess.run(["git", "clone", "--branch", branch, repo_url, tmp_dir], check=True)
        git_dir = os.path.join(tmp_dir, ".git")
        if os.path.exists(git_dir):
            shutil.rmtree(git_dir, ignore_errors=True)
        output_file = os.path.join(os.getcwd(), "parsed_repo_scaffold.json")
        subprocess.run(["node", "app/ts_parser.js", tmp_dir, output_file], check=True)
        with open(output_file, "r", encoding="utf-8") as f:
            parsed_json = json.load(f)
        ingest_parsed_scaffold(parsed_json)
        return parsed_json
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)

if st.button("ðŸ“¥ Pull & Ingest Repo", key="btn_pull_ingest_repo"):
    if not repo_url.strip():
        st.warning("Please provide a valid repo URL")
    else:
        try:
            with st.spinner("Cloning repo and parsing with TS-Morph..."):
                parsed_json = pull_and_ingest_repo(repo_url, branch)
            st.success(f"âœ… Repo scaffold ingested successfully: {len(parsed_json.get('modules', []))} modules")
            # Make the agentic generator use this same repository by default
            try:
                st.session_state.framework_repo_path = repo_url.strip()
                st.info("Framework Repo Path updated to the provided Git URL for agentic script generation.")
            except Exception:
                pass
        except subprocess.CalledProcessError as e:
            st.error(f"Git/Parser command failed: {e}")
        except Exception as ex:
            st.error(f"Unexpected error: {ex}")

st.title("Test Artifact Recorder & Ingest")

# ========================== Agentic AI Test Script Generator ==========================
st.header("Agentic AI Test Script Generator - Conversational Mode")

# Initialize conversation
if "conversation" not in st.session_state:
    st.session_state.conversation = []

# ------------------- User Input -------------------
user_input = st.text_input("Ask or type scenario / feedback:", key="chat_input")
uploaded_file = st.file_uploader(
    "Optional: Upload scenario/template/test data file", 
    type=["json","xlsx","xls","xlsm","csv","txt","doc","docx","testdata"],
    key="chat_file_uploader"
)

def flatten_file_keywords(uploaded_file):
    """Extract keywords from uploaded file"""
    keywords = []
    if uploaded_file:
        fname = uploaded_file.name
        ext = fname.split(".")[-1].lower()
        if ext in ["xlsx", "xls", "xlsm"]:
            df = load_excel_template(uploaded_file, dtype=str)
            keywords = df.fillna("").astype(str).to_numpy().flatten().tolist()
        elif ext == "json":
            data = json.load(uploaded_file)
            if isinstance(data, dict):
                keywords = list(data.values())
            elif isinstance(data, list):
                keywords = data
        elif ext == "csv":
            content = uploaded_file.getvalue().decode(errors="ignore")
            keywords = [row for row in content.splitlines() if row.strip()]
        elif ext == "testdata":
            content = uploaded_file.getvalue().decode(errors="ignore")
            keywords = [row for row in content.splitlines() if row.strip()]
        else:
            # txt/doc/docx fallback
            content = uploaded_file.getvalue().decode(errors="ignore")
            keywords = content.splitlines()
    return keywords

def stream_ai_response(content: str, lang="typescript"):
    placeholder = st.empty()
    chunk_size = 50
    for i in range(0, len(content), chunk_size):
        chunk = content[:i+chunk_size]
        if lang:
            placeholder.code(chunk, language=lang)
        else:
            placeholder.markdown(chunk)
        time.sleep(0.02)
    return placeholder

def detect_intent(msg: str):
    msg_lower = msg.lower()
    msg_normalized = re.sub(r"\btrail\b", "trial", msg_lower)
    source_text = msg_normalized
    if any(
        k in source_text
        for k in [
            "trial run",
            "run trial",
            "execute test",
            "run test",
            "try the script",
            "execute script",
            "run the flow",
            "playwright run",
            "execute scenario",
        ]
    ):
        return "trial_run"
    if any(k in source_text for k in ["compare", "difference", "diff"]) and "flow" in source_text:
        return "compare_flow"
    if any(k in source_text for k in ["latest flow", "recent flow", "show latest", "newest flow"]):
        return "latest_flow"
    if any(k in source_text for k in ["push", "commit", "publish to git", "push to github"]):
        return "push_code"
    if any(
        k in source_text
        for k in [
            "generate script",
            "create test script",
            "test script",
            "script preview",
            "automation script",
            "playwright script",
            "new script",
            "build script",
        ]
    ):
        return "agentic_script"
    if any(k in source_text for k in ["generate draft", "draft", "flow", "scenario"]):
        return "draft"
    if any(k in source_text for k in ["apply feedback", "modify", "change"]):
        return "feedback"
    if any(k in source_text for k in ["preview script", "show script", "script"]):
        return "agentic_script"
    return "unknown"

def search_existing_script(keywords):
    """Check existing framework / Vector DB for full actionable script"""
    artifacts = db.query(keywords, top_k=5)
    docs_list = artifacts.get("documents", []) if isinstance(artifacts, dict) else artifacts
    for doc in docs_list:
        content = doc.get("content", "") if isinstance(doc, dict) else str(doc)
        if keywords in content or keywords.replace(" ", "") in content:
            return content
    return None


def normalize_remote_repo_input(repo_input: str) -> Tuple[str, Optional[str]]:
    cleaned = repo_input.replace("\\", "/").strip()
    cleaned = cleaned.replace("https:/", "https://").replace("http:/", "http://")
    branch_in_url = None

    if cleaned.startswith("git@"):
        return cleaned, branch_in_url

    if "://" not in cleaned and cleaned.startswith("github.com"):
        cleaned = f"https://{cleaned}"

    if cleaned.startswith("http") and "/tree/" in cleaned:
        base, remainder = cleaned.split("/tree/", 1)
        branch_in_url = remainder.split("/", 1)[0]
        cleaned = base

    if cleaned.endswith("/"):
        cleaned = cleaned[:-1]

    if cleaned.startswith("http") and not cleaned.endswith(".git"):
        cleaned = f"{cleaned}.git"

    return cleaned, branch_in_url


def resolve_framework_repo(repo_input: str, branch: str) -> Tuple[Path, str]:
    repo_input = (repo_input or "").strip()
    if not repo_input:
        raise ValueError("Framework repo path is empty.")

    desired_branch = branch.strip() if branch else ""

    if any(repo_input.startswith(prefix) for prefix in ("http://", "https://", "git@")) or "github.com" in repo_input:
        clone_url, branch_in_url = normalize_remote_repo_input(repo_input)
        branch_to_use = branch_in_url or desired_branch
        slug_source = clone_url + (f"#{branch_to_use}" if branch_to_use else "")
        slug = hashlib.sha1(slug_source.encode("utf-8")).hexdigest()[:12]
        target = (FRAMEWORK_CLONE_BASE / slug).resolve()

        if not target.exists():
            target.parent.mkdir(parents=True, exist_ok=True)
            subprocess.run(["git", "clone", clone_url, str(target)], check=True)

        subprocess.run(["git", "-C", str(target), "fetch", "origin"], check=True)
        if branch_to_use:
            subprocess.run(["git", "-C", str(target), "checkout", branch_to_use], check=True)
            subprocess.run(["git", "-C", str(target), "pull", "origin", branch_to_use], check=True)
        else:
            current_branch = subprocess.check_output(
                ["git", "-C", str(target), "rev-parse", "--abbrev-ref", "HEAD"],
                text=True,
                encoding="utf-8",
                errors="replace",
            ).strip()
            branch_to_use = current_branch

        return target, branch_to_use

    path = Path(repo_input).expanduser()
    if not path.is_absolute():
        path = Path(os.getcwd()) / path
    path = path.resolve()
    if not path.exists():
        raise FileNotFoundError(f"Framework repo not found: {path}")
    if not desired_branch and (path / ".git").exists():
        try:
            desired_branch = subprocess.check_output(
                ["git", "-C", str(path), "rev-parse", "--abbrev-ref", "HEAD"],
                text=True,
                encoding="utf-8",
                errors="replace",
            ).strip()
        except subprocess.CalledProcessError:
            desired_branch = "main"

    return path, desired_branch or "main"


def handle_agentic_message(message: str, intent: str) -> List[Dict[str, str]]:
    if intent == "latest_flow":
        return present_latest_flow_details()

    state = st.session_state.agentic_state
    repo_path = st.session_state.framework_repo_path.strip()

    if not repo_path:
        return [
            {
                "role": "assistant",
                "content": "Please set the Framework Repo Path in the sidebar before requesting script generation.",
                "type": "text",
            }
        ]

    try:
        resolved_path, active_branch = resolve_framework_repo(repo_path, st.session_state.framework_branch)
    except (FileNotFoundError, ValueError, subprocess.CalledProcessError) as exc:
        return [{"role": "assistant", "content": f"{exc}", "type": "text"}]

    st.session_state.framework_branch = active_branch
    st.session_state.resolved_framework_path = str(resolved_path)

    framework = FrameworkProfile.from_root(resolved_path)

    agent = agentic_engine
    data_upload_messages = persist_uploaded_data_file(framework)
    responses: List[Dict[str, str]] = list(data_upload_messages)

    status = state.get("status")

    if status == "awaiting-datasheet":
        defaults = state.get("pending_datasheet_defaults") or derive_default_datasheet_fields(state.get("scenario", ""))
        state["pending_datasheet_defaults"] = defaults

        parsed = parse_datasheet_message(message, defaults)
        if parsed:
            pending_ids = state.get("pending_test_ids") or []
            if not pending_ids:
                pending_ids = [state.get("scenario", "").strip()] if state.get("scenario") else []

            update_messages: List[Dict[str, str]] = []
            for test_id in pending_ids:
                if not test_id:
                    continue
                update_info = update_test_manager_entry(
                    framework,
                    test_id,
                    execute_value="Yes",
                    create_if_missing=True,
                    datasheet=parsed["datasheet"],
                    reference_id=parsed["reference_id"],
                    id_name=parsed["id_name"],
                )
                if update_info:
                    register_config_update(state, update_messages, update_info)
                else:
                    update_messages.append(
                        {
                            "role": "assistant",
                            "content": f"Could not update testmanager.xlsx for '{test_id}'. Please update it manually.",
                            "type": "text",
                        }
                    )

            responses.extend(update_messages)
            state["datasheet_values"] = parsed
            state["status"] = "script-ready"
            state["awaiting_datasheet"] = False
            state["pending_test_ids"] = []

            confirmation_message = (
                "Datasheet mapping recorded: "
                f"{parsed['datasheet']} / {parsed['reference_id']} / {parsed['id_name']}. "
                "Upload the workbook to the data/ folder if you haven't already. "
                "You can now run `trial run` or provide additional feedback."
            )
            responses.append({"role": "assistant", "content": confirmation_message, "type": "text"})

        else:
            responses.append(
                {
                    "role": "assistant",
                    "content": (
                        "Datasheet mapping is optional. You can reply now with "
                        "`datasheet <file> reference <id> idname <column>` (or `use defaults`) to update testmanager.xlsx. "
                        "Otherwise, you may proceed with `trial run` or continue providing feedback. "
                        "If the workbook is missing at runtime, the script will prompt you to upload it."
                    ),
                    "type": "text",
                }
            )
            state["status"] = "script-ready"
            state["awaiting_datasheet"] = False
            state["pending_test_ids"] = []
        return responses

    if intent == "trial_run":
        headed = "headed" in message.lower()
        scenario_hint = re.sub(r"\b(trial|run|execute|flow|scenario|in|mode|headed)\b", " ", message, flags=re.IGNORECASE).strip()
        trial_responses = execute_trial_run(state, framework, scenario_hint, headed=headed)
        return responses + trial_responses

    if intent == "compare_flow":
        return responses + compare_repo_and_refined(message, framework)

    if intent == "push_code":
        if state.get("status") in {"ready-for-push", "script-ready"}:
            return responses + execute_push_to_github(state, framework)
        return responses + [
            {
                "role": "assistant",
                "content": "No script payload is ready to push. Confirm the preview and generate the code before pushing.",
                "type": "text",
            }
        ]

    if not state["active"] or state.get("status") in {"idle", "complete"} or intent == "agentic_script" and state.get("status") == "idle":
        st.session_state.agentic_state = initialise_agentic_state()
        state = st.session_state.agentic_state
        state["active"] = True
        state["scenario"] = message
        state["status"] = "preview-awaiting"
        state["feedback"] = []
        explicit_refined = bool(re.search(r"\brefined\b", message, re.IGNORECASE))
        # Prefer existing framework assets if available before invoking LLM context collection
        existing_assets = agent.find_existing_framework_assets(message, framework)
        repo_contents: List[str] = []
        if existing_assets:
            state["existing_files"] = [str(asset["path"].relative_to(framework.root)) for asset in existing_assets]
            files_list = []
            collected_ids: List[str] = []
            for asset in existing_assets:
                rel = asset["path"].relative_to(framework.root)
                files_list.append(str(rel))
                try:
                    content = asset["path"].read_text(encoding="utf-8")
                except UnicodeDecodeError:
                    content = "(Binary or non-UTF-8 content omitted)"
                repo_contents.append(content)
                responses.append({"role": "assistant", "content": f"// {rel}\n{content}", "type": "script"})
                if asset["path"].suffix == ".ts":
                    collected_ids.extend(extract_test_ids_from_content(content))
            if collected_ids:
                enable_tests_for_ids(framework, collected_ids, state, responses)
            else:
                update_info = update_test_manager_entry(framework, state["scenario"], execute_value="Yes")
                if update_info:
                    register_config_update(state, responses, update_info)
                else:
                    responses.append(
                        {
                            "role": "assistant",
                            "content": "Could not locate testmanager.xlsx to enable this scenario; please review the repo manually.",
                            "type": "text",
                        }
                    )
            summary = "Existing framework files located (repo scan):\n" + "\n".join(f"- {path}" for path in files_list)
            responses.append({"role": "assistant", "content": summary, "type": "text"})
        context = agent.gather_context(message)
        state["context"] = context
        vector_steps = context.get("vector_steps") or []
        if vector_steps:
            responses.append(
                {
                    "role": "assistant",
                    "content": format_refined_steps(vector_steps),
                    "type": "text",
                }
            )
            diff_summary = format_refined_diff(vector_steps, repo_contents)
            if diff_summary:
                responses.append({"role": "assistant", "content": diff_summary, "type": "text"})
        # If there is zero usable context (no repo assets, no recorder flow, no vector-derived steps), do not hallucinate.
        has_flow = context.get("flow_available")
        has_enriched = bool(context.get("enriched_steps"))
        if not existing_assets and not has_flow and not has_enriched:
            responses.append({
                "role": "assistant",
                "content": "No relevant data found in framework repo, recorder flows, or vector DB for this scenario. Please record a flow or ingest context, then try again.",
                "type": "text",
            })
            state["status"] = "complete"
            state["active"] = False
            return responses
        if existing_assets and not vector_steps and not has_flow and not explicit_refined:
            state["status"] = "complete"
            state["active"] = False
            return responses

        preview = agent.generate_preview(message, framework, context)
        state["preview"] = preview
        responses.append({"role": "assistant", "content": preview, "type": "preview"})
        if not context.get("flow_available"):
            responses.append(
                {
                    "role": "assistant",
                    "content": "Recorder flow not found. Preview steps are derived from Jira/documentation/repository context.",
                    "type": "text",
                }
            )
        responses.append(
            {
                "role": "assistant",
                "content": "Review the preview steps. Reply with feedback to refine them or say 'confirm' to generate the full script.",
                "type": "text",
            }
        )
        return responses

    status = state.get("status")

    if status == "preview-awaiting":
        if interpret_confirmation(message):
            context_snapshot = state.get("context") or {}
            vector_steps = context_snapshot.get("vector_steps") or []
            has_flow = context_snapshot.get("flow_available")
            if not vector_steps and not has_flow:
                existing_assets = agent.find_existing_framework_assets(state["scenario"], framework)
            else:
                existing_assets = []
            if existing_assets:
                state["status"] = "complete"
                state["active"] = False
                state["existing_files"] = [
                    str(asset["path"].relative_to(framework.root)) for asset in existing_assets
                ]
                files_list = []
                collected_ids_confirm: List[str] = []
                for asset in existing_assets:
                    rel = asset["path"].relative_to(framework.root)
                    files_list.append(str(rel))
                    try:
                        content = asset["path"].read_text(encoding="utf-8")
                    except UnicodeDecodeError:
                        content = "(Binary or non-UTF-8 content omitted)"
                    responses.append({"role": "assistant", "content": f"// {rel}\n{content}", "type": "script"})
                    if asset["path"].suffix == ".ts":
                        collected_ids_confirm.extend(extract_test_ids_from_content(content))
                if collected_ids_confirm:
                    enable_tests_for_ids(framework, collected_ids_confirm, state, responses)
                else:
                    update_info = update_test_manager_entry(framework, state["scenario"], execute_value="Yes")
                    if update_info:
                        register_config_update(state, responses, update_info)
                    else:
                        responses.append({"role": "assistant", "content": "Could not locate testmanager.xlsx to enable this scenario; please review the repo manually.", "type": "text"})
                summary = "Existing framework files located:\n" + "\n".join(f"- {path}" for path in files_list)
                responses.append({"role": "assistant", "content": summary, "type": "text"})
                return responses
            try:
                payload = agent.generate_script_payload(state["scenario"], framework, state["preview"])
            except Exception as exc:  # noqa: BLE001
                return [
                    {
                        "role": "assistant",
                        "content": f"Failed to generate script: {exc}",
                        "type": "text",
                    }
                ]

            state["payload"] = payload
            state["status"] = "awaiting-datasheet"
            state["written_files"] = []

            generated_test_ids: List[str] = []

            for section, files in payload.items():
                for file_obj in files:
                    display = f"// {file_obj['path']}\n{file_obj['content']}"
                    responses.append({"role": "assistant", "content": display, "type": "script"})
                    if section == "tests":
                        generated_test_ids.extend(extract_test_ids_from_content(file_obj.get('content', '')))

            pending_ids = generated_test_ids or [state["scenario"]]
            state["pending_test_ids"] = pending_ids
            defaults = derive_default_datasheet_fields(pending_ids[0] if pending_ids else state["scenario"])
            state["pending_datasheet_defaults"] = defaults
            state["awaiting_datasheet"] = True
            state["datasheet_values"] = None

            instruction_message = (
                "Script generated. Before running a trial, provide the datasheet mapping so I can update "
                "testmanager.xlsx. Upload the Excel file to the framework's data/ folder if needed, then reply with "
                "`datasheet <file> reference <id> idname <column>` or say `use defaults`.\n"
                f"Suggested defaults â†’ DatasheetName: {defaults['datasheet']}, "
                f"ReferenceID: {defaults['reference_id']}, IDName: {defaults['id_name']}."
            )
            responses.append({"role": "assistant", "content": instruction_message, "type": "text"})
            return responses

        if interpret_push(message):
            return [
                {
                    "role": "assistant",
                    "content": "Preview must be confirmed before pushing. Please confirm the steps first.",
                    "type": "text",
                }
            ]

        state.setdefault("feedback", []).append(message)
        previous_context = state.get("context", {})
        refreshed_context = agent.gather_context(state["scenario"])
        state["context"] = refreshed_context
        if refreshed_context.get("flow_available") and not (previous_context or {}).get("flow_available"):
            responses.append(
                {
                    "role": "assistant",
                    "content": "A newer recorder flow was found and will be incorporated into the updated preview.",
                    "type": "text",
                }
            )
        state.setdefault("feedback", []).append(message)
        refined = agent.refine_preview(
            state["scenario"],
            framework,
            state["preview"],
            message,
            refreshed_context,
        )
        state["preview"] = refined
        responses.append({"role": "assistant", "content": refined, "type": "preview"})
        responses.append(
            {
                "role": "assistant",
                "content": "Preview updated. Reply 'confirm' when ready or continue sharing feedback.",
                "type": "text",
            }
        )
        return responses

    if status == "script-ready":
        if interpret_push(message):
            return execute_push_to_github(state, framework)

        if interpret_feedback(message):
            prev_context = state.get("context", {})
            refreshed_context = agent.gather_context(state["scenario"])
            state["context"] = refreshed_context
            if refreshed_context.get("flow_available") and not (prev_context or {}).get("flow_available"):
                responses.append(
                    {
                        "role": "assistant",
                        "content": "A recorder flow is now available and will be used for the updated preview.",
                        "type": "text",
                    }
                )
            state.setdefault("feedback", []).append(message)
            refined = agent.refine_preview(
                state["scenario"],
                framework,
                state["preview"],
                message,
                refreshed_context,
            )
            state["preview"] = refined
            state["status"] = "preview-awaiting"
            state["payload"] = {}
            state["written_files"] = []
            state["pending_test_ids"] = []
            state["pending_datasheet_defaults"] = None
            state["datasheet_values"] = None
            state["awaiting_datasheet"] = False
            responses.append({"role": "assistant", "content": refined, "type": "preview"})
            responses.append(
                {
                    "role": "assistant",
                    "content": "Script discarded. Review the updated preview and confirm when ready.",
                    "type": "text",
                }
            )
            return responses

        responses.append(
            {
                "role": "assistant",
                "content": "Script is ready. Reply 'push' to persist or provide feedback to adjust the flow.",
                "type": "text",
            }
        )
        return responses

    responses.append(
        {
            "role": "assistant",
            "content": "Agentic session complete. Start a new request for another script.",
            "type": "text",
        }
    )
    return responses

# ------------------- Process User Input -------------------
if st.button("Send") and user_input.strip():
    if uploaded_file:
        ext = Path(uploaded_file.name).suffix.lower()
        if ext in {".xlsx", ".xls", ".xlsm", ".csv", ".testdata"}:
            st.session_state["pending_data_upload"] = {
                "name": Path(uploaded_file.name).name,
                "bytes": uploaded_file.getvalue(),
            }
    file_keywords = flatten_file_keywords(uploaded_file)
    combined_keywords = " ".join([user_input] + file_keywords).strip()

    st.session_state.conversation.append({"role": "user", "content": user_input})

    state = st.session_state.agentic_state
    intent = detect_intent(user_input)

    target_intent = intent if intent != "unknown" else "agentic_script"
    if state.get("active") and intent == "unknown":
        target_intent = "agentic_script"
    replies = handle_agentic_message(combined_keywords or user_input, target_intent)

    for reply in replies:
        st.session_state.conversation.append(reply)

# ------------------- Display Conversation (latest first) -------------------
for msg in reversed(st.session_state.conversation):
    if msg["role"] == "user":
        st.markdown(f"**You:** {msg['content']}")
    else:
        msg_type = msg.get("type")
        if msg_type == "script":
            st.markdown("**AI (Script):**")
            st.code(msg["content"], language="typescript")
        elif msg_type in {"preview", "draft"}:
            st.markdown(f"**AI (Preview):**\n{msg['content']}")
        else:
            st.markdown(f"**AI:** {msg['content']}")

